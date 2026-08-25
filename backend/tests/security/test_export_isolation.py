"""HTTP-level cross-tenant isolation for GET /api/export — proves the
export's batched query is scoped by RLS, not just individual expense reads,
and that this holds under every filter combination and both formats.
"""

import io
from datetime import date
from decimal import Decimal

import openpyxl
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.expense import Expense
from app.db.models.user import User


def _auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _seed_victim(owner_session: AsyncSession, **overrides) -> None:
    victim = User(email="victim-export@example.com", password_hash="not-a-real-hash")
    owner_session.add(victim)
    await owner_session.flush()

    defaults = dict(
        vendor="Victim's Vendor",
        category="Software",
        currency="AUD",
        total=Decimal("777.00"),
        expense_date=date.today(),
        status="confirmed",
    )
    defaults.update(overrides)
    owner_session.add(Expense(user_id=victim.id, **defaults))
    await owner_session.commit()


async def test_csv_export_never_includes_another_users_rows_under_no_filter(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    await _seed_victim(owner_session)
    token_a = await signup_user("attacker-export-csv@example.com")

    resp = await client.get("/api/export", headers=_auth_headers(token_a))

    assert resp.status_code == 200
    body = resp.content.decode("utf-8-sig")
    assert "Victim's Vendor" not in body
    assert "777.00" not in body
    # Headers-only: just the column header row.
    assert body.strip().count("\n") == 0


async def test_csv_export_never_includes_another_users_rows_under_any_filter(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    await _seed_victim(owner_session)
    token_a = await signup_user("attacker-export-csv-filtered@example.com")

    for params in (
        {"category": "Software"},
        {"status": "confirmed"},
        {"q": "victim"},
        {"date_from": "2000-01-01"},
    ):
        resp = await client.get("/api/export", params=params, headers=_auth_headers(token_a))
        assert resp.status_code == 200
        body = resp.content.decode("utf-8-sig")
        assert "Victim's Vendor" not in body
        assert "777.00" not in body


async def test_xlsx_export_never_includes_another_users_rows(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    await _seed_victim(owner_session)
    token_a = await signup_user("attacker-export-xlsx@example.com")

    resp = await client.get(
        "/api/export", params={"format": "xlsx"}, headers=_auth_headers(token_a)
    )

    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    expenses_sheet = workbook["Expenses"]
    rows = list(expenses_sheet.iter_rows(values_only=True))
    assert len(rows) == 1  # header only
    line_items_sheet = workbook["Line Items"]
    assert len(list(line_items_sheet.iter_rows(values_only=True))) == 1  # header only


async def test_owner_can_still_export_their_own_rows_alongside_a_victim(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    """Sanity check: isolation isn't just blocking everyone."""
    await _seed_victim(owner_session)
    token_a = await signup_user("owner-export-sanity@example.com")
    me = await client.get("/api/auth/me", headers=_auth_headers(token_a))
    user_id = me.json()["id"]

    owner_session.add(
        Expense(
            user_id=user_id,
            vendor="My Own Vendor",
            category="Software",
            currency="AUD",
            total=Decimal("42.00"),
            expense_date=date.today(),
            status="confirmed",
        )
    )
    await owner_session.commit()

    resp = await client.get("/api/export", headers=_auth_headers(token_a))

    assert resp.status_code == 200
    body = resp.content.decode("utf-8-sig")
    assert "My Own Vendor" in body
    assert "Victim's Vendor" not in body
