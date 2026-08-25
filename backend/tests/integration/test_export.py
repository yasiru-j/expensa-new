"""GET /api/export — exercises CSV and XLSX generation against the real
Postgres database: filter parity with GET /api/expenses, currency handling,
CSV-injection escaping, BOM/encoding, empty results, and line items.
"""

import csv
import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.expense import Expense
from app.db.models.line_item import LineItem


def _auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _get_user_id(client: AsyncClient, token: str) -> str:
    resp = await client.get("/api/auth/me", headers=_auth_headers(token))
    return resp.json()["id"]


def _seed(owner_session: AsyncSession, user_id: str, **overrides) -> Expense:
    defaults = dict(
        vendor="Test Vendor",
        category="Meals",
        currency="AUD",
        subtotal=Decimal("10.00"),
        tax=Decimal("0.00"),
        total=Decimal("10.00"),
        expense_date=date.today(),
        status="ready",
        payment_method="Card",
    )
    defaults.update(overrides)
    expense = Expense(user_id=user_id, **defaults)
    owner_session.add(expense)
    return expense


def _parse_csv(content: bytes) -> list[list[str]]:
    assert content.startswith(b"\xef\xbb\xbf"), "CSV must start with a UTF-8 BOM"
    text = content[3:].decode("utf-8")
    return list(csv.reader(io.StringIO(text)))


async def test_export_csv_rows_match_filtered_list_for_same_params(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-parity@example.com")
    user_id = await _get_user_id(client, token)

    _seed(owner_session, user_id, vendor="Corner Cafe", category="Meals")
    _seed(owner_session, user_id, vendor="Some Software", category="Software")
    _seed(owner_session, user_id, vendor="Corner Cafe Too", category="Meals", status="confirmed")
    await owner_session.commit()

    params = {"category": "Meals"}
    listing = await client.get("/api/expenses", params=params, headers=_auth_headers(token))
    export = await client.get("/api/export", params=params, headers=_auth_headers(token))

    assert listing.status_code == 200
    assert export.status_code == 200

    listed_vendors = sorted(item["vendor"] for item in listing.json()["items"])
    rows = _parse_csv(export.content)
    header, data_rows = rows[0], rows[1 : 1 + listing.json()["total"]]
    assert header == [
        "id",
        "vendor",
        "vendor_tax_id",
        "expense_date",
        "subtotal",
        "tax",
        "total",
        "currency",
        "category",
        "payment_method",
        "status",
        "created_at",
    ]
    exported_vendors = sorted(row[1] for row in data_rows)
    assert exported_vendors == listed_vendors == ["Corner Cafe", "Corner Cafe Too"]


async def test_export_csv_respects_date_and_vendor_search_filters(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-datesearch@example.com")
    user_id = await _get_user_id(client, token)

    today = date.today()
    _seed(owner_session, user_id, vendor="In Range", expense_date=today)
    _seed(owner_session, user_id, vendor="Too Early", expense_date=today - timedelta(days=30))
    await owner_session.commit()

    params = {
        "date_from": (today - timedelta(days=1)).isoformat(),
        "date_to": (today + timedelta(days=1)).isoformat(),
        "q": "in range",
    }
    export = await client.get("/api/export", params=params, headers=_auth_headers(token))

    rows = _parse_csv(export.content)
    data_rows = [r for r in rows[1:] if r and r[0]]
    assert len(data_rows) == 1
    assert data_rows[0][1] == "In Range"


async def test_export_paginates_internally_across_many_rows(
    client: AsyncClient, signup_user, owner_session: AsyncSession, monkeypatch
) -> None:
    """Forces a tiny internal batch size so a 5-row export spans multiple
    batches, proving the streaming pagination doesn't skip or duplicate rows."""
    monkeypatch.setattr("app.api.export.BATCH_SIZE", 2)
    token = await signup_user("export-batches@example.com")
    user_id = await _get_user_id(client, token)

    for i in range(5):
        _seed(owner_session, user_id, vendor=f"Vendor {i}", total=Decimal(f"{i}.00"))
    await owner_session.commit()

    export = await client.get("/api/export", headers=_auth_headers(token))
    rows = _parse_csv(export.content)
    data_rows = [r for r in rows[1:] if r and r[0] and r[3] != "TOTAL (AUD)"]
    assert len(data_rows) == 5
    assert sorted(r[1] for r in data_rows) == [f"Vendor {i}" for i in range(5)]


async def test_export_csv_currency_column_present_and_totals_grouped_not_summed(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-currency@example.com")
    user_id = await _get_user_id(client, token)

    _seed(owner_session, user_id, vendor="AUD One", currency="AUD", total=Decimal("10.00"))
    _seed(owner_session, user_id, vendor="AUD Two", currency="AUD", total=Decimal("5.50"))
    _seed(owner_session, user_id, vendor="USD One", currency="USD", total=Decimal("20.00"))
    await owner_session.commit()

    export = await client.get("/api/export", headers=_auth_headers(token))
    rows = _parse_csv(export.content)

    header = rows[0]
    currency_idx = header.index("currency")
    total_idx = header.index("total")
    data_rows = [r for r in rows[1:] if r and r[1] and not r[3].startswith("TOTAL")]
    for row in data_rows:
        assert row[currency_idx] in ("AUD", "USD")

    totals_rows = [r for r in rows[1:] if r and r[3].startswith("TOTAL")]
    totals_by_currency = {r[currency_idx]: r[total_idx] for r in totals_rows}
    assert totals_by_currency == {"AUD": "15.50", "USD": "20.00"}


async def test_export_decimal_precision_has_no_float_artifacts(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-precision@example.com")
    user_id = await _get_user_id(client, token)

    _seed(
        owner_session,
        user_id,
        vendor="Precise Vendor",
        subtotal=Decimal("12.34"),
        tax=Decimal("1.20"),
        total=Decimal("13.54"),
    )
    await owner_session.commit()

    export = await client.get("/api/export", headers=_auth_headers(token))
    rows = _parse_csv(export.content)
    header = rows[0]
    row = next(r for r in rows[1:] if r and r[1] == "Precise Vendor")

    assert row[header.index("subtotal")] == "12.34"
    assert row[header.index("tax")] == "1.20"
    assert row[header.index("total")] == "13.54"


async def test_export_csv_injection_is_escaped_on_malicious_vendor_names(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-injection@example.com")
    user_id = await _get_user_id(client, token)

    malicious_vendors = [
        "=cmd|'/c calc'!A1",
        "+1+1",
        "-2+3",
        "@SUM(A1:A9)",
        "Safe Vendor",
    ]
    for vendor in malicious_vendors:
        _seed(owner_session, user_id, vendor=vendor, payment_method='=HYPERLINK("evil")')
    await owner_session.commit()

    export = await client.get("/api/export", headers=_auth_headers(token))
    rows = _parse_csv(export.content)
    header = rows[0]
    data_rows = [r for r in rows[1:] if r and r[1]]

    vendors = {r[header.index("vendor")] for r in data_rows}
    assert vendors == {
        "'=cmd|'/c calc'!A1",
        "'+1+1",
        "'-2+3",
        "'@SUM(A1:A9)",
        "Safe Vendor",
    }
    payment_methods = {r[header.index("payment_method")] for r in data_rows}
    assert '\'=HYPERLINK("evil")' in payment_methods
    assert '=HYPERLINK("evil")' not in payment_methods


async def test_export_csv_utf8_bom_preserves_non_ascii_vendor_names(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-utf8@example.com")
    user_id = await _get_user_id(client, token)

    _seed(owner_session, user_id, vendor="Café Résumé 日本語")
    await owner_session.commit()

    export = await client.get("/api/export", headers=_auth_headers(token))
    assert export.content.startswith(b"\xef\xbb\xbf")

    rows = _parse_csv(export.content)
    vendors = {r[1] for r in rows[1:] if r and r[1]}
    assert "Café Résumé 日本語" in vendors


async def test_export_empty_filtered_result_yields_valid_headers_only_csv(
    client: AsyncClient, signup_user
) -> None:
    token = await signup_user("export-empty@example.com")

    export = await client.get(
        "/api/export", params={"category": "Software"}, headers=_auth_headers(token)
    )

    assert export.status_code == 200
    rows = _parse_csv(export.content)
    assert len(rows) == 1
    assert rows[0][0] == "id"


async def test_export_csv_content_type_and_filename(client: AsyncClient, signup_user) -> None:
    token = await signup_user("export-headers@example.com")

    export = await client.get("/api/export", headers=_auth_headers(token))

    assert export.status_code == 200
    assert export.headers["content-type"].startswith("text/csv")
    disposition = export.headers["content-disposition"]
    assert "attachment" in disposition
    assert disposition.split("filename=")[1].strip('"').endswith(".csv")
    assert "expensa-expenses-" in disposition


async def test_export_xlsx_content_type_and_filename(client: AsyncClient, signup_user) -> None:
    token = await signup_user("export-xlsx-headers@example.com")

    export = await client.get(
        "/api/export", params={"format": "xlsx"}, headers=_auth_headers(token)
    )

    assert export.status_code == 200
    assert export.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = export.headers["content-disposition"]
    assert disposition.split("filename=")[1].strip('"').endswith(".xlsx")


async def test_export_xlsx_expenses_sheet_matches_filtered_list_and_includes_line_items(
    client: AsyncClient, signup_user, owner_session: AsyncSession
) -> None:
    token = await signup_user("export-xlsx-lineitems@example.com")
    user_id = await _get_user_id(client, token)

    expense = _seed(owner_session, user_id, vendor="Receipt With Items", total=Decimal("30.00"))
    await owner_session.flush()
    owner_session.add(
        LineItem(
            expense_id=expense.id,
            description="Widget",
            quantity=Decimal("2"),
            unit_price=Decimal("10.00"),
            amount=Decimal("20.00"),
        )
    )
    owner_session.add(
        LineItem(
            expense_id=expense.id,
            description="Gadget",
            quantity=Decimal("1"),
            unit_price=Decimal("10.00"),
            amount=Decimal("10.00"),
        )
    )
    _seed(owner_session, user_id, vendor="Receipt Without Items", total=Decimal("5.00"))
    await owner_session.commit()

    export = await client.get(
        "/api/export", params={"format": "xlsx"}, headers=_auth_headers(token)
    )
    assert export.status_code == 200

    workbook = openpyxl.load_workbook(io.BytesIO(export.content))
    assert workbook.sheetnames == ["Expenses", "Line Items"]

    expenses_sheet = workbook["Expenses"]
    expense_rows = list(expenses_sheet.iter_rows(values_only=True))
    header = expense_rows[0]
    vendors = {row[header.index("vendor")] for row in expense_rows[1:] if row[0]}
    assert vendors == {"Receipt With Items", "Receipt Without Items"}

    totals_rows = [row for row in expense_rows[1:] if row[3] and str(row[3]).startswith("TOTAL")]
    assert len(totals_rows) == 1
    assert totals_rows[0][header.index("currency")] == "AUD"
    assert str(totals_rows[0][header.index("total")]) == "35.00"

    line_items_sheet = workbook["Line Items"]
    line_item_rows = list(line_items_sheet.iter_rows(values_only=True))
    assert line_item_rows[0] == ("expense_id", "description", "quantity", "unit_price", "amount")
    descriptions = {row[1] for row in line_item_rows[1:]}
    assert descriptions == {"Widget", "Gadget"}
    for row in line_item_rows[1:]:
        assert row[0] == str(expense.id)
