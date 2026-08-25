"""GET /api/export — streams the caller's expenses (owner-scoped via RLS) as
CSV or Excel, filtered identically to GET /api/expenses via the shared
app.api.expense_filters.build_expense_conditions.

Design notes (see README's Export section for the user-facing version):

- Format is chosen via a `format` query param (csv|xlsx), not Accept-header
  content negotiation — every other filter on this API is already a query
  param, and a query param makes the export linkable/downloadable directly
  from a plain <a href> without any client-side content-negotiation code.
- CSV is genuinely streamed row-by-row, batched off the database
  (BATCH_SIZE at a time) inside one transaction, so the full result set is
  never held in memory as Python objects. XLSX cannot be streamed the same
  way — the .xlsx container is a zip archive that isn't valid until fully
  written — so it's built into an in-memory buffer with openpyxl's
  write_only mode (which still avoids holding a full worksheet object graph)
  and returned as one chunk. This is documented rather than overclaiming
  full XLSX streaming.
- Line items are excluded from the flat CSV (they don't fit a "one row per
  expense" shape without duplicating expense rows) but included in XLSX as a
  second "Line Items" sheet, keyed by expense_id, so no data is silently
  dropped from the richer format.
- A totals-by-currency block is appended after the data rows in both
  formats, computed by summing over the DECIMAL total column as rows are
  written (never converted to float) and never summed across currencies —
  the same principle as the Phase 4 dashboard's per-currency aggregates.
- Vendor / vendor_tax_id / payment_method are free text an attacker fully
  controls (a receipt's OCR'd vendor name). A value starting with =, +, -,
  or @ is prefixed with a leading apostrophe, which is exactly what Excel
  does when a user manually types a literal value that looks like a
  formula — it defuses formula/DDE injection while leaving the visible text
  otherwise unchanged. This mitigates CSV formula injection in the CSV
  export; XLSX cell values written by openpyxl are typed as plain strings
  (not formulas) so they aren't executable there either, but the same
  sanitization is applied for consistency between the two formats.
"""

import csv
import io
import uuid
from collections.abc import AsyncIterator
from datetime import UTC, datetime
from datetime import date as date_type
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select

from app.api.expense_filters import (
    CATEGORY_PATTERN,
    SORT_OPTIONS,
    SORT_PATTERN,
    STATUS_PATTERN,
    build_expense_conditions,
)
from app.core.deps import get_current_user_id, user_scoped_session
from app.db.maintenance import sweep_stale_processing_rows
from app.db.models.expense import Expense
from app.db.models.line_item import LineItem

router = APIRouter(prefix="/api/export", tags=["export"])

BATCH_SIZE = 500
CSV_INJECTION_PREFIXES = ("=", "+", "-", "@")

EXPENSE_COLUMNS = [
    "id",
    "vendor",
    "vendor_tax_id",
    "expense_date",
    "subtotal",
    "tax",
    "total",
    "currency",
    "category",
    "payment_method",
    "status",
    "created_at",
]
LINE_ITEM_COLUMNS = ["expense_id", "description", "quantity", "unit_price", "amount"]


def _sanitize_text(value: str | None) -> str:
    value = value or ""
    if value and value[0] in CSV_INJECTION_PREFIXES:
        return "'" + value
    return value


def _decimal_str(value: Decimal | None) -> str:
    return "" if value is None else format(value, "f")


def _expense_row(expense: Expense) -> list[str]:
    return [
        str(expense.id),
        _sanitize_text(expense.vendor),
        _sanitize_text(expense.vendor_tax_id),
        expense.expense_date.isoformat() if expense.expense_date else "",
        _decimal_str(expense.subtotal),
        _decimal_str(expense.tax),
        _decimal_str(expense.total),
        expense.currency or "",
        expense.category or "",
        _sanitize_text(expense.payment_method),
        expense.status,
        expense.created_at.isoformat(),
    ]


def _line_item_row(line_item: LineItem) -> list[str]:
    return [
        str(line_item.expense_id),
        _sanitize_text(line_item.description),
        _decimal_str(line_item.quantity),
        _decimal_str(line_item.unit_price),
        _decimal_str(line_item.amount),
    ]


def _totals_row(currency: str, total: Decimal) -> list[str]:
    row = [""] * len(EXPENSE_COLUMNS)
    row[3] = f"TOTAL ({currency})"
    row[6] = format(total, "f")
    row[7] = currency
    return row


def _add_to_totals(totals: dict[str, Decimal], expense: Expense) -> None:
    if expense.total is not None:
        key = expense.currency or "UNKNOWN"
        totals[key] = totals.get(key, Decimal("0")) + expense.total


async def _fetch_batches(
    user_id: uuid.UUID, conditions: list, sort: str
) -> AsyncIterator[list[Expense]]:
    """Yields the matching expenses in BATCH_SIZE chunks, all within one
    transaction scoped to user_id — never loading the full result set at
    once. `Expense.id` is appended as a tiebreaker so paginating with
    OFFSET/LIMIT can't skip or duplicate rows on ties in the primary sort key.
    """
    order_clause = SORT_OPTIONS[sort]
    async with user_scoped_session(user_id) as session:
        await sweep_stale_processing_rows(session)
        offset = 0
        while True:
            query = (
                select(Expense).order_by(order_clause, Expense.id).offset(offset).limit(BATCH_SIZE)
            )
            for condition in conditions:
                query = query.where(condition)
            batch = (await session.execute(query)).scalars().all()
            if not batch:
                return
            yield batch
            if len(batch) < BATCH_SIZE:
                return
            offset += BATCH_SIZE


async def _stream_csv(user_id: uuid.UUID, conditions: list, sort: str) -> AsyncIterator[bytes]:
    # UTF-8 BOM so Excel (which otherwise guesses the system codepage) opens
    # non-ASCII vendor names correctly instead of mangling them.
    yield b"\xef\xbb\xbf"

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPENSE_COLUMNS)
    yield buffer.getvalue().encode("utf-8")
    buffer.seek(0)
    buffer.truncate(0)

    totals: dict[str, Decimal] = {}
    async for batch in _fetch_batches(user_id, conditions, sort):
        for expense in batch:
            writer.writerow(_expense_row(expense))
            _add_to_totals(totals, expense)
        yield buffer.getvalue().encode("utf-8")
        buffer.seek(0)
        buffer.truncate(0)

    if totals:
        for currency in sorted(totals):
            writer.writerow(_totals_row(currency, totals[currency]))
        yield buffer.getvalue().encode("utf-8")


async def _build_xlsx(user_id: uuid.UUID, conditions: list, sort: str) -> bytes:
    workbook = Workbook(write_only=True)
    expenses_sheet = workbook.create_sheet("Expenses")
    expenses_sheet.append(EXPENSE_COLUMNS)
    line_items_sheet = workbook.create_sheet("Line Items")
    line_items_sheet.append(LINE_ITEM_COLUMNS)

    totals: dict[str, Decimal] = {}
    async with user_scoped_session(user_id) as session:
        await sweep_stale_processing_rows(session)
        order_clause = SORT_OPTIONS[sort]
        offset = 0
        while True:
            query = (
                select(Expense).order_by(order_clause, Expense.id).offset(offset).limit(BATCH_SIZE)
            )
            for condition in conditions:
                query = query.where(condition)
            batch = (await session.execute(query)).scalars().all()
            if not batch:
                break

            expense_ids = []
            for expense in batch:
                expenses_sheet.append(_expense_row(expense))
                _add_to_totals(totals, expense)
                expense_ids.append(expense.id)

            line_items = (
                (
                    await session.execute(
                        select(LineItem)
                        .where(LineItem.expense_id.in_(expense_ids))
                        .order_by(LineItem.expense_id)
                    )
                )
                .scalars()
                .all()
            )
            for line_item in line_items:
                line_items_sheet.append(_line_item_row(line_item))

            if len(batch) < BATCH_SIZE:
                break
            offset += BATCH_SIZE

    for currency in sorted(totals):
        expenses_sheet.append(_totals_row(currency, totals[currency]))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.get("")
async def export_expenses(
    user_id: uuid.UUID = Depends(get_current_user_id),
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    sort: str = Query("date_desc", pattern=SORT_PATTERN),
    status_filter: str | None = Query(None, alias="status", pattern=STATUS_PATTERN),
    date_from: date_type | None = Query(None),
    date_to: date_type | None = Query(None),
    category: str | None = Query(None, pattern=CATEGORY_PATTERN),
    q: str | None = Query(
        None,
        min_length=1,
        max_length=200,
        description="Case-insensitive, partial-match vendor search",
    ),
) -> StreamingResponse:
    conditions = build_expense_conditions(
        status_filter=status_filter,
        date_from=date_from,
        date_to=date_to,
        category=category,
        q=q,
    )
    stamp = datetime.now(UTC).strftime("%Y%m%d")

    if format == "xlsx":
        content = await _build_xlsx(user_id, conditions, sort)
        filename = f"expensa-expenses-{stamp}.xlsx"
        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    filename = f"expensa-expenses-{stamp}.csv"
    return StreamingResponse(
        _stream_csv(user_id, conditions, sort),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
